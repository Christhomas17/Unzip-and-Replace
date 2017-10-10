# -*- coding: utf-8 -*-
"""
Created on Tue Sep 19 08:06:25 2017

@author: builduser
"""

import zipfile
import os
import shutil

from openpyxl import load_workbook,Workbook

import openpyxl
from OpenPyXLHelperFunctions import data_from_range, data_to_range

reelSrcRange = 'C9:L30'
reelDestRange = 'B5'

cwd = os.getcwd()

def unzip_all(cwd):
    os.chdir(cwd)
    fileList = []
    for file in os.listdir():
        try:
            name,extension = file.split('.')
        except:
            name = file
            extension = 'apple'
            
        if extension == 'zip':
            name = file.split('_')[0]
            fileList.append(name + '.xlsm')
            with zipfile.ZipFile(file,'r') as filename:
                filename.extractall()
        

    return(fileList)
                


def copy_and_rename_all(fileList):
    filez = []
    for file in fileList:
        src = 'Math.xlsx'
        dest = 'DGE - ' + file.split('.')[0] + '.xlsx'
        
        filez.append(dest)
        
        shutil.copy(src,dest)
        
    return(filez)
        
def copy_and_paste(fileList,filez):
    if len(fileList) == len(filez):
        for index in range(len(fileList)):
            wb = load_workbook(fileList[index], data_only = True)
            data = data_from_range(reelSrcRange,'Strips',wb)
            wb.close()
            
            wb = load_workbook(filez[index], data_only= True)
            data_to_range(data,reelDestRange,wb.get_sheet_by_name('Table'),wb)
            
            wb.save(filez[index])
    
            
fileList = unzip_all(cwd)
print('unzipped')
filez = copy_and_rename_all(fileList)
print('renamed')
copy_and_paste(fileList,filez)
